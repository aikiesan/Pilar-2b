"""
IBGE PPM — Pesquisa da Pecuária Municipal: herds, animal products, aquaculture.

Closes the largest verified data gap in the national expansion: `municipalities`
has cattle_/swine_/poultry_biogas_m3_year columns that are ZERO for all 4,926
non-São-Paulo municipalities, and the ABIOVE deliverable supplies nothing here
(it is crop/LULC-only). PPM is the yearly, municipal, authoritative herd census.

Raw snapshots (SIDRA "Ano x variável" exports, one workbook per table, each
covering the WHOLE series — hence no per-year subfolder under data/raw/):

    data/raw/ibge_ppm/TABELA_3939_PPM_2008A2024.xlsx   herds        2008-2024
    data/raw/ibge_ppm/TABELA_74_2008A2024.xlsx         products     2008-2024
    data/raw/ibge_ppm/TABELA_3940_2008A2024.xlsx       aquaculture  2013-2024

    python -m ingest.runner run ibge_ppm --year 2024

SIDRA LAYOUT (identical in all three): row 3 = Nível/Cód./Município,
row 4 = year (**merged cells — forward-fill required**), row 5 = variable,
data from row 6. `Cód.` is a real 7-digit IBGE code, so there is no name join
and none of the municipality-matching traps that plague the other sources.

VALUE CODES ARE NOT INTERCHANGEABLE
    '-'          -> 0     (measured as zero)
    '..' / '...' -> NULL  (not available / not surveyed)
Collapsing '..' to 0 would invent zero herds where IBGE simply did not survey,
which then propagates into biogas potential as "this municipality has none".

THE TRAPS (each is a real double-count, unit error, or silent schema break)

1. SUBSET COLUMNS. `Suíno - matrizes de suínos` ⊂ `Suíno - total`, and
   `Galináceos - galinhas` ⊂ `Galináceos - total`. Summing all ten herd types
   double-counts. They are loaded under distinct names and MUST NOT be added to
   any total; validate() asserts the containment so the trap cannot go unnoticed.

2. `Total` IS NOT A TOTAL — IT SUMS INCOMPATIBLE UNITS. In table 74 the columns
   are Mil litros (milk), Mil dúzias (eggs) and Quilogramas (honey). Their
   "Total" is arithmetic nonsense, which is exactly why IBGE renders it as '..'
   for every row. Never ingest it. This is why municipality_timeseries.unit is
   NOT NULL.

3. NESTED AGGREGATE + AN EMPTY 2024 REPLACEMENT IN TABLE 3940. For 2013-2023
   `Peixes` is a group HEADER over 18 species columns and carries no data ('..'
   in all 4,091 rows). In 2024 IBGE flattened the hierarchy to
   `Peixes (Quilogramas)` (27 columns that year vs 26) — but that column is
   *also* empty: 3,479 non-null cells, every one of them zero. So neither
   variant is usable and the fish total is ALWAYS summed from the 18 species
   (2024 sum = 724,853 t, consistent with Brazil's ~600-900 kt/yr of farmed
   fish). _fish_method_gate guards that decision: if IBGE ever populates the
   column, it fails and tells us to switch, instead of leaving us summing out
   of habit.

4. UNITS LIVE IN THE HEADER TEXT: 'Leite (Mil litros)', 'Mel de abelha
   (Quilogramas)', 'Alevinos (Milheiros)', and even 'Outros produtos (... )
   (Nenhuma)' — a genuinely unitless column. Units are therefore declared
   per-variable in VARIABLES below, never inferred at write time.

COVERAGE: 3939 carries 5,567 of the 5,571 municipalities. The four absentees are
legitimate and allow-listed, not silent gaps — see MISSING_ALLOWLIST.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from ingest.contract import IngestContext, SourceSpec
from ingest.gates import GateResult, run_standard_battery

RAW_SUBDIR = "ibge_ppm"

HERDS_FILE = "TABELA_3939_PPM_2008A2024.xlsx"
PRODUCTS_FILE = "TABELA_74_2008A2024.xlsx"
AQUACULTURE_FILE = "TABELA_3940_2008A2024.xlsx"

HERDS_SHEET = "Tabela"
PRODUCTS_SHEET = "Produção de origem animal"
AQUACULTURE_SHEET = "Produção da aquicultura"

# SIDRA column label -> (our variable name, unit). Anything not listed is dropped
# on purpose, including every `Total` pseudo-column (trap 2).
VARIABLES: dict[str, tuple[str, str]] = {
    # -- table 3939, herds (unit: head) -------------------------------------
    "Bovino": ("bovino", "head"),
    "Suíno - total": ("suino_total", "head"),
    "Galináceos - total": ("galinaceos_total", "head"),
    "Bubalino": ("bubalino", "head"),
    "Caprino": ("caprino", "head"),
    "Ovino": ("ovino", "head"),
    "Equino": ("equino", "head"),
    "Codornas": ("codornas", "head"),
    # Subsets of the two totals above (trap 1). Kept for completeness because
    # matrizes/laying-hen counts matter for manure modelling, but they are never
    # part of a herd total.
    "Suíno - matrizes de suínos": ("suino_matrizes", "head"),
    "Galináceos - galinhas": ("galinaceos_galinhas", "head"),
    # -- table 74, animal products -------------------------------------------
    "Leite (Mil litros)": ("leite", "mil_litros"),
    "Ovos de galinha (Mil dúzias)": ("ovos_galinha", "mil_duzias"),
    "Ovos de codorna (Mil dúzias)": ("ovos_codorna", "mil_duzias"),
    "Mel de abelha (Quilogramas)": ("mel", "kg"),
    # -- table 3940, aquaculture (2013+) --------------------------------------
    "Camarão (Quilogramas)": ("aquicultura_camarao", "kg"),
    "Ostras, vieiras e mexilhões (Quilogramas)": ("aquicultura_moluscos", "kg"),
}

# Fish species that make up `Peixes` for 2013-2023 (trap 3). Summed into
# aquicultura_peixes; in 2024 the direct column below is used instead.
FISH_SPECIES = (
    "Carpa (Quilogramas)",
    "Curimatã, curimbatá (Quilogramas)",
    "Dourado (Quilogramas)",
    "Jatuarana, piabanha e piracanjuba (Quilogramas)",
    "Lambari (Quilogramas)",
    "Matrinxã (Quilogramas)",
    "Pacu e patinga (Quilogramas)",
    "Piau, piapara, piauçu, piava (Quilogramas)",
    "Pintado, cachara, cachapira e pintachara, surubim (Quilogramas)",
    "Pirapitinga (Quilogramas)",
    "Pirarucu (Quilogramas)",
    "Tambacu, tambatinga (Quilogramas)",
    "Tambaqui (Quilogramas)",
    "Tilápia (Quilogramas)",
    "Traíra e trairão (Quilogramas)",
    "Truta (Quilogramas)",
    "Tucunaré (Quilogramas)",
    "Outros peixes (Quilogramas)",
)
FISH_DIRECT_COLUMN = "Peixes (Quilogramas)"  # 2024 only
FISH_VARIABLE = "aquicultura_peixes"

# Unit per emitted variable, including the derived fish total.
UNIT_BY_VARIABLE: dict[str, str] = {v: u for v, u in VARIABLES.values()}
UNIT_BY_VARIABLE[FISH_VARIABLE] = "kg"

# Municipalities with no PPM row, verified against the spine (2026-07-16). None
# is a parsing failure — IBGE does not survey livestock where there is none:
MISSING_ALLOWLIST = (
    "3300258",  # Arraial do Cabo/RJ — coastal/fishing, no livestock surveyed
    "3500600",  # Águas de São Pedro/SP — smallest municipality in Brazil (3.38 km²)
    "3548807",  # São Caetano do Sul/SP — dense industrial ABC, no rural area
    "5101837",  # Boa Esperança do Norte/MT — created after 2017; absent from the whole series
)

# IBGE's OWN published national figures, per year -> {variable: (total, tolerance_pct)}.
# These are what make the aggregation gate real: they are an independent statement
# by the source about what the municipal rows must add up to, so reproducing them
# proves the parse (year forward-fill, value codes, column mapping) is correct.
#
# Source: IBGE Agência de Notícias, "Valor da produção da pecuária e da aquicultura
# chega a R$ 132,8 bilhões em 2024" (PPM 2024 release).
#
# Tolerance follows the PRECISION OF THE PUBLISHED FIGURE, not our confidence:
# a number printed as "238,2 milhões" carries 4 significant digits and supports a
# tight test; one printed as "1,6 bilhão" carries 2 and cannot support one.
PUBLISHED_NATIONAL: dict[int, dict[str, tuple[float, float]]] = {
    2024: {
        # "o efetivo bovino atingiu 238,2 milhões de cabeças"
        "bovino": (238_200_000, 0.5),
        # "o efetivo (...) de galinhas a 277,5 milhões, ambos recordes"
        "galinaceos_galinhas": (277_500_000, 0.5),
        # NOT tested: galináceos total is published only as "1,6 bilhão" — 2
        # significant digits, i.e. ±3% inherent. Our 1,581,215,552 rounds to
        # exactly that, but asserting ±1% against it would test IBGE's rounding
        # rather than our parse, and would fail for the wrong reason.
    },
}

# Known inconsistencies in IBGE's own published data: (year, ibge_code, subset,
# total) triples where the source publishes a subset larger than its parent.
# Each entry is investigated and justified below — the gate tolerates exactly
# these and still fails on anything new, so this stays a ledger of known source
# defects rather than a way to mute the check.
#
# Values are stored AS PUBLISHED. We record the defect rather than repair it:
# inventing a plausible total would put a number in the database that IBGE never
# said, and the neighbouring years are right there for anyone who needs to judge.
# A full sweep of 2008-2024 (2026-07-16) found exactly FOUR violations across
# ~94,000 municipality-years — the series is otherwise internally consistent.
# All four are small-magnitude reporting slips in IBGE's published data.
KNOWN_SUBSET_ANOMALIES: set[tuple[int, str, str, str]] = {
    # Japeri/RJ 2017: total published as 0 while galinhas = 120. Neighbouring
    # years are consistent (2016 total=2,335; 2018 total=122), so the zero is a
    # reporting error, not a real collapse. 120 head vs ~1.4 billion nationally.
    (2017, "3302270", "galinaceos_galinhas", "galinaceos_total"),
    # Primavera/PE 2020: galinhas 400 vs total 250 (excess 150 head).
    (2020, "2611408", "galinaceos_galinhas", "galinaceos_total"),
    # Itaquara/BA 2021: suíno matrizes 19 vs total 10 (excess 9 head).
    (2021, "2916708", "suino_matrizes", "suino_total"),
    # Andradina/SP 2023: galinhas 8,900 vs total 8,853 (excess 47 head, 0.5%).
    (2023, "3502101", "galinaceos_galinhas", "galinaceos_total"),
}

SPEC = SourceSpec(
    source_id="ibge_ppm",
    name="IBGE PPM — herds, animal products and aquaculture by municipality",
    key_column="ibge_code",
    # Only the three biogas-primary herds are required: they exist in every year
    # 2008-2024. Products and aquaculture are optional (3940 starts in 2013 and
    # covers 4,091 municipalities), so requiring them would fail the schema gate
    # on legitimate years.
    required_columns={
        "ibge_code": "str",
        "bovino": "float",
        "suino_total": "float",
        "galinaceos_total": "float",
    },
    # Bounds derived from the observed maxima across 2008-2024 (measured, not
    # invented), each with headroom. The maxima cross-check against known
    # Brazilian agricultural geography, which is what makes them citable:
    #   bovino          2,522,608  São Félix do Xingu/PA 2022 (largest herd in BR)
    #   suino_total     1,242,843  Toledo/PR 2015 (swine hub)
    #   galinaceos_total 20,020,000 Cascavel/PR 2021 (poultry hub)
    #   codornas         4,000,000  Bastos/SP 2014 (quail capital of BR)
    value_bounds={
        "bovino": (0, 4_000_000),
        "suino_total": (0, 2_000_000),
        "galinaceos_total": (0, 40_000_000),
        "codornas": (0, 8_000_000),
        "bubalino": (0, 500_000),
        "caprino": (0, 1_500_000),
        "ovino": (0, 1_500_000),
        "equino": (0, 100_000),
    },
    missing_allowlist=MISSING_ALLOWLIST,
    aggregation_tolerance_pct=1.0,
)


def _parse_value(raw: object) -> float | None:
    """SIDRA cell -> float. '-' is a measured zero; '..'/'...'/'X' are absences.

    Conflating the two would fabricate zero herds wherever IBGE did not survey.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text == "-":
        return 0.0
    if text in ("..", "...", "X", ""):
        return None
    try:
        return float(text.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def _read_sidra_year(path: Path, sheet: str, year: int) -> dict[str, dict[str, float | None]]:
    """Read one SIDRA workbook and return {ibge_code: {sidra_label: value}} for `year`.

    Row 4 holds the year in a merged cell, so it is forward-filled across the
    block; row 5 holds the variable label.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        years: list[str | None] = []
        last: str | None = None
        for col in range(4, ws.max_column + 1):
            value = ws.cell(row=4, column=col).value
            if value:
                last = str(value).strip()
            years.append(last)
        labels = [ws.cell(row=5, column=col).value for col in range(4, ws.max_column + 1)]

        wanted = [i for i, y in enumerate(years) if y == str(year)]
        # The workbook simply has no columns for this year (e.g. table 3940 starts
        # in 2013). Return empty rather than a dict of empty rows, so callers can
        # tell "not covered by this table" from "covered but all values absent".
        if not wanted:
            return {}
        out: dict[str, dict[str, float | None]] = {}
        for row in ws.iter_rows(min_row=6, values_only=True):
            # The footer ("Fonte: IBGE — Pesquisa da Pecuária Municipal") also
            # lands in column 0; only 'MU' rows are municipalities.
            if not row or row[0] != "MU" or not row[1]:
                continue
            code = str(row[1]).strip()
            cells: dict[str, float | None] = {}
            for i in wanted:
                label = labels[i]
                if not label:
                    continue
                pos = 3 + i
                cells[str(label).strip()] = _parse_value(row[pos] if pos < len(row) else None)
            out[code] = cells
        return out
    finally:
        wb.close()


def fetch(year: int, raw_dir: Path) -> Path:
    """Verify the snapshots exist; never download silently (contract rule).

    Note the deliberate absence of a per-year subdirectory: each SIDRA workbook
    ships the entire 2008-2024 series, so one snapshot serves every --year and
    load() slices the year out.
    """
    base = Path(raw_dir) / RAW_SUBDIR
    missing = [f for f in (HERDS_FILE, PRODUCTS_FILE, AQUACULTURE_FILE) if not (base / f).exists()]
    if missing:
        raise FileNotFoundError(
            f"IBGE PPM snapshots missing from {base}: {missing}\n"
            "Export tables 3939 (Efetivo dos rebanhos), 74 (Produção de origem animal) "
            "and 3940 (Produção da aquicultura) from https://sidra.ibge.gov.br as "
            "'Ano x variável' municipal workbooks, place them at the paths above, and "
            "record the retrieval date in docs/data/METADATA.json."
        )
    return base


def load(year: int, raw_dir: Path) -> pd.DataFrame:
    """One WIDE row per municipality for `year`, columns = curated variables.

    Wide (not long) because the standard battery is built for it: schema_gate
    checks ibge_code uniqueness and coverage_gate counts one row per
    municipality per UF. promote() melts this to the long shape that
    municipality_timeseries stores.
    """
    base = fetch(year, raw_dir)

    herds = _read_sidra_year(base / HERDS_FILE, HERDS_SHEET, year)
    products = _read_sidra_year(base / PRODUCTS_FILE, PRODUCTS_SHEET, year)
    aquaculture = _read_sidra_year(base / AQUACULTURE_FILE, AQUACULTURE_SHEET, year)

    if not herds:
        raise ValueError(f"table 3939 has no data for {year} — check the snapshot's year range")

    # The row set is 3939's municipalities; the other two left-join onto it.
    # Anchoring on one table keeps the coverage gate deterministic instead of
    # drifting with each table's own reporting universe (74 has 5,569 rows,
    # 3940 only 4,091 — aquaculture simply does not exist everywhere).
    records: list[dict[str, object]] = []
    for code in sorted(herds):
        row: dict[str, object] = {"ibge_code": code}
        merged = {**herds[code], **products.get(code, {}), **aquaculture.get(code, {})}

        for label, (variable, _unit) in VARIABLES.items():
            if label in merged:
                row[variable] = merged[label]

        # Fish total is ALWAYS summed from the species columns — never read from
        # `Peixes`/`Peixes (Quilogramas)`, neither of which carries data (trap 3).
        species = [merged.get(s) for s in FISH_SPECIES if merged.get(s) is not None]
        row[FISH_VARIABLE] = float(sum(species)) if species else None

        records.append(row)

    df = pd.DataFrame.from_records(records)
    # The three primary herds must be numeric for the schema gate even in the
    # (impossible-in-practice) case where a whole column came back empty.
    for column in ("bovino", "suino_total", "galinaceos_total"):
        if column not in df.columns:
            df[column] = pd.NA
        df[column] = pd.to_numeric(df[column], errors="coerce")
    df["ibge_code"] = df["ibge_code"].astype(str)
    return df


def _subset_gate(df: pd.DataFrame, subset: str, total: str, year: int) -> GateResult:
    """A subset column may never exceed its parent (trap 1).

    This is the guard that makes the double-count trap loud: if a future export
    renames or reorders these columns, `matrizes > total` fires here instead of
    silently inflating a herd sum downstream. It has already earned its keep by
    catching a genuine IBGE reporting error (see KNOWN_SUBSET_ANOMALIES).
    """
    name = f"subset[{subset}<={total}]"
    if subset not in df.columns or total not in df.columns:
        return GateResult(name, True, "column absent this year; skipped")

    pair = df[["ibge_code", subset, total]].copy()
    pair[subset] = pd.to_numeric(pair[subset], errors="coerce")
    pair[total] = pd.to_numeric(pair[total], errors="coerce")
    pair = pair.dropna(subset=[subset, total])

    violations = pair[pair[subset] > pair[total]]
    unknown = [
        r
        for _, r in violations.iterrows()
        if (year, str(r["ibge_code"]), subset, total) not in KNOWN_SUBSET_ANOMALIES
    ]
    known = len(violations) - len(unknown)

    if unknown:
        worst = unknown[0]
        return GateResult(
            name,
            False,
            f"{len(unknown)} UNDOCUMENTED rows where {subset} > {total} "
            f"(e.g. {worst['ibge_code']}: {worst[subset]:,.0f} > {worst[total]:,.0f}) — "
            "investigate, then add to KNOWN_SUBSET_ANOMALIES with a justification",
        )
    detail = f"{len(pair)} rows satisfy {subset} <= {total}"
    if known:
        detail += f" ({known} known source anomaly tolerated)"
    return GateResult(name, True, detail)


def _fish_method_gate(year: int, raw_dir: Path, df: pd.DataFrame) -> GateResult:
    """Guard the decision to always sum species rather than read IBGE's own total.

    Neither `Peixes` (the 2013-2023 group header) nor `Peixes (Quilogramas)` (its
    2024 replacement) carries data: measured 2026-07-16, the 2024 column has
    3,479 non-null cells of which every one is zero. Summing the 18 species is
    therefore the only method that yields a fish total at all.

    This gate exists to catch the day that stops being true. If IBGE ever
    populates the column, it fires — telling us to read the published total
    instead of continuing to sum, rather than silently keeping a now-inferior
    method.
    """
    base = Path(raw_dir) / RAW_SUBDIR
    raw = _read_sidra_year(base / AQUACULTURE_FILE, AQUACULTURE_SHEET, year)
    if not raw:
        return GateResult("fish-total-method", True, f"{year}: table 3940 has no data (pre-2013)")

    published = sum(v for cells in raw.values() if (v := cells.get(FISH_DIRECT_COLUMN)) is not None)
    summed = float(pd.to_numeric(df.get(FISH_VARIABLE), errors="coerce").fillna(0).sum())

    if published > 0:
        return GateResult(
            "fish-total-method",
            False,
            f"IBGE now publishes a non-empty '{FISH_DIRECT_COLUMN}' ({published:,.0f} kg) — "
            f"stop summing species (currently {summed:,.0f} kg) and read it directly",
        )
    return GateResult(
        "fish-total-method",
        True,
        f"'{FISH_DIRECT_COLUMN}' still empty as expected; fish total summed from "
        f"{len(FISH_SPECIES)} species = {summed:,.0f} kg",
    )


def _published_totals_gate(df: pd.DataFrame, year: int) -> GateResult:
    """Reproduce IBGE's own published national figures for every variable whose
    published form is precise enough to test (see PUBLISHED_NATIONAL)."""
    expected = PUBLISHED_NATIONAL.get(year)
    if not expected:
        return GateResult(
            "published-totals",
            True,
            f"no IBGE published national figures recorded for {year}; add them to "
            "PUBLISHED_NATIONAL to make this gate bite",
        )
    problems, checked = [], []
    for variable, (published, tol_pct) in expected.items():
        if variable not in df.columns:
            problems.append(f"{variable}: column absent")
            continue
        got = float(pd.to_numeric(df[variable], errors="coerce").fillna(0).sum())
        drift = abs(got - published) / published
        if drift * 100 > tol_pct:
            problems.append(
                f"{variable}: {got:,.0f} vs published {published:,.0f} ({drift:.2%} > {tol_pct}%)"
            )
        else:
            checked.append(f"{variable} {drift:.2%}")
    if problems:
        return GateResult("published-totals", False, "; ".join(problems))
    return GateResult(
        "published-totals", True, f"matches IBGE's published {year} figures: {', '.join(checked)}"
    )


def validate(
    df: pd.DataFrame,
    ctx: IngestContext | None = None,
    raw_dir: Path | None = None,
) -> tuple[list[GateResult], IngestContext]:
    """Standard battery plus the source-specific containment guards."""
    if ctx is None:
        ctx = IngestContext(spec=SPEC, year=0)

    # Feed the standard aggregation gate IBGE's own published national figure for
    # the headline herd, so it tests something real instead of failing for want of
    # input. Callers may override by setting published_totals themselves.
    if not ctx.published_totals and ctx.year in PUBLISHED_NATIONAL:
        ctx.published_totals = {"BR": PUBLISHED_NATIONAL[ctx.year]["bovino"][0]}
        ctx.totals_column = "bovino"

    results = run_standard_battery(df, ctx)
    results.append(_published_totals_gate(df, ctx.year))
    results.append(_subset_gate(df, "suino_matrizes", "suino_total", ctx.year))
    results.append(_subset_gate(df, "galinaceos_galinhas", "galinaceos_total", ctx.year))
    if raw_dir is not None:
        results.append(_fish_method_gate(ctx.year, raw_dir, df))
    return results, ctx


def promote(df: pd.DataFrame, conn) -> None:
    """Wide frame -> long municipality_timeseries rows, in ONE transaction.

    Idempotent by construction: the table's UNIQUE (ibge_code, year, source_id,
    variable) turns a re-run into an UPDATE rather than a duplicate — the
    structural defence against the kind of double-load that inflated ABIOVE's
    TerraClass table 2.89x.
    """
    raise NotImplementedError(
        "Promotion runs via backend/scripts/promote_ibge_ppm.py against the real DB — "
        "database writes stay a reviewed, manual step (ingest/README.md)."
    )
