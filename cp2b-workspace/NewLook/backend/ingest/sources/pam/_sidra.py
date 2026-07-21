"""Shared SIDRA reader for the two PAM tables (1612 temporárias, 1613 permanentes).

Both tables ship in the same "Ano x produto" layout the PPM tables use, so the
row geometry below is deliberately identical to ingest/sources/ibge_ppm — one
parser shape for every SIDRA export we take.

    row 3   Nível | Cód. | Município | <block header>
    row 4   year, in MERGED cells -> forward-fill required
    row 5   product name, repeating per year block
    row 6+  data; column A = Nível, B = Cód. (real 7-digit IBGE), C = name

WHAT DIFFERS FROM PPM, AND WHY IT MATTERS

1. ONE SHEET PER VARIABLE. PAM splits "Área plantada", "Área colhida" and
   "Quantidade produzida (Toneladas)" into separate sheets of the same workbook,
   where PPM had a single sheet. We read *Quantidade produzida* for the biomass
   columns; área colhida is read separately as an independent cross-check.
   Sheet names are truncated by Excel at 31 chars ("Quantidade produzida
   (Tonela..."), so they are matched by prefix, never by equality.

2. THE SERIES IS SPLIT ACROSS SEVERAL WORKBOOKS, WITH OVERLAPS. 1613 ships both
   ..._2016_A_2020 and ..._2020_A_2023 — 2020 appears in both. IBGE revises
   figures in later releases, so when a year is available from more than one
   workbook we take it from the one with the LATEST start year (newest vintage)
   and record which file it came from. Silently taking whichever sorted first
   would make the load depend on filename order.

VALUE CODES ARE NOT INTERCHANGEABLE (identical rule to PPM)
    '-'            -> 0     measured as zero: the crop is not grown here
    '..' / '...'   -> NULL  not surveyed / not available
    'X'            -> NULL  value withheld to protect the informant

'X' is the one PAM leans on heavily (≈19.8k cells, concentrated in café and
laranja): the municipality *does* produce, but IBGE suppresses the number. It is
therefore an absence, not a zero — writing 0 would assert the municipality grows
no coffee, which is precisely false. It reaches the map as no_data.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

# Data starts at column D (1-indexed 4); A/B/C are Nível, Cód., Município.
_FIRST_DATA_COL = 4
_HEADER_YEAR_ROW = 4
_HEADER_PRODUCT_ROW = 5
_FIRST_DATA_ROW = 6

PRODUCED_SHEET_PREFIX = "Quantidade produzida"
HARVESTED_SHEET_PREFIX = "Área colhida"

# TABELA_1612_2024_A_2021.xlsx -> (1612, 2024, 2021); ranges appear in both orders.
_FILENAME_RE = re.compile(r"TABELA_(\d{4})_(\d{4})(?:_A_(\d{4}))?", re.IGNORECASE)


def parse_value(raw: object) -> float | None:
    """SIDRA cell -> float. '-' is a measured zero; '..'/'...'/'X' are absences."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text == "-":
        return 0.0
    if text in ("..", "...", "X", "x", ""):
        return None
    try:
        return float(text.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def _sheet_by_prefix(workbook: openpyxl.Workbook, prefix: str) -> str:
    """Match a sheet by prefix — Excel truncates long sheet names at 31 chars."""
    folded = prefix.casefold()
    for name in workbook.sheetnames:
        if name.casefold().startswith(folded[: min(len(folded), 25)]):
            return name
    raise ValueError(f"no sheet starting with {prefix!r}; have {workbook.sheetnames}")


def workbooks_for_table(raw_dir: Path, table: int) -> list[Path]:
    return sorted(p for p in raw_dir.glob("TABELA_*.xlsx") if f"TABELA_{table}_" in p.name.upper())


def _file_year_span(path: Path) -> tuple[int, int] | None:
    match = _FILENAME_RE.search(path.name)
    if not match:
        return None
    years = [int(y) for y in match.groups()[1:] if y]
    return (min(years), max(years)) if years else None


def select_workbook(raw_dir: Path, table: int, year: int) -> Path:
    """Pick the workbook supplying `year`, preferring the newest vintage.

    Where ranges overlap (1613 ships 2016_A_2020 and 2020_A_2023) the later
    release carries IBGE's revised figures, so it wins.
    """
    candidates = []
    for path in workbooks_for_table(raw_dir, table):
        span = _file_year_span(path)
        if span and span[0] <= year <= span[1]:
            candidates.append((span[0], path))
    if not candidates:
        raise FileNotFoundError(
            f"no TABELA_{table} workbook in {raw_dir} covers {year}; "
            f"found {[p.name for p in workbooks_for_table(raw_dir, table)]}"
        )
    return max(candidates)[1]


def read_year(path: Path, year: int, sheet_prefix: str = PRODUCED_SHEET_PREFIX) -> dict:
    """{ibge_code: {product: value|None}} for one year of one workbook.

    Row 4 carries the year only in the first cell of each merged block, so it is
    forward-filled across the row before the year is matched.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[_sheet_by_prefix(workbook, sheet_prefix)]
        rows = list(
            worksheet.iter_rows(
                min_row=_HEADER_YEAR_ROW, max_row=_HEADER_PRODUCT_ROW, values_only=True
            )
        )
        year_row, product_row = rows[0], rows[1]

        # Forward-fill the merged year cells, then keep the columns for `year`.
        columns: list[int] = []
        products: list[str] = []
        current: str | None = None
        for index in range(_FIRST_DATA_COL - 1, len(product_row)):
            cell = year_row[index] if index < len(year_row) else None
            if cell is not None and str(cell).strip():
                current = str(cell).strip()
            product = product_row[index]
            if current == str(year) and product:
                columns.append(index)
                products.append(str(product).strip())

        if not columns:
            return {}

        out: dict[str, dict[str, float | None]] = {}
        for row in worksheet.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True):
            if not row or row[0] != "MU":
                continue  # skip UF/region aggregate rows and trailing notes
            code = str(row[1]).strip() if row[1] is not None else ""
            if len(code) != 7 or not code.isdigit():
                continue
            out[code] = {
                product: parse_value(row[index] if index < len(row) else None)
                for index, product in zip(columns, products)
            }
        return out
    finally:
        workbook.close()
