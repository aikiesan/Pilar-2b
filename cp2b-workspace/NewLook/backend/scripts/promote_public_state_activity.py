"""Build and promote the reviewed SP + MG livestock/urban activity snapshot.

The committed CSV is compact verification evidence, not a replacement for the
primary sources. Rebuild it explicitly with ``--build-snapshot`` from:

* IBGE SIDRA table 3939 (PPM 2024 municipal herds);
* IBGE SIDRA table 4709 (Censo 2022 municipal population); and
* SNIS-RS 2022 ``Planilha_Informacoes_RS_2022.xlsx`` (CO111/CO115/CO119).

Promotion is UF-scoped and idempotent. PPM values remain measured head counts
in ``municipality_timeseries``; the API converts them to estimated manure mass
with the canonical per-head factors. SNIS waste values retain the source unit,
tonnes/year. Missing SNIS reports remain NULL and use the explicitly estimated
population fallback already implemented by the map API.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

PUBLIC_STATE_COUNTS = {"31": 853, "35": 645}
PUBLIC_MUNICIPALITIES = sum(PUBLIC_STATE_COUNTS.values())
YEAR_PPM = 2024
YEAR_CENSO = 2022
YEAR_SNIS = 2022

PPM_URL = (
    "https://apisidra.ibge.gov.br/values/t/3939/n6/in%20n3%20{uf}/v/all/p/2024/c79/all?formato=json"
)
CENSO_URL = "https://apisidra.ibge.gov.br/values/t/4709/n6/in%20n3%20{uf}/v/all/p/all?formato=json"
SNIS_URL = (
    "https://www.gov.br/cidades/pt-br/acesso-a-informacao/acoes-e-programas/"
    "saneamento/snis/produtos-do-snis/diagnosticos/"
    "Planilha_RS_2022_atualizado_29112024.zip"
)
DEFAULT_SNAPSHOT = (
    Path(__file__).resolve().parents[1] / "data" / "pilot" / "public_states_activity_2022_2024.csv"
)

HERD_VARIABLES = {
    "Bovino": "cattle_head_2024",
    "Suíno - total": "swine_head_2024",
    "Galináceos - total": "poultry_head_2024",
}

FIELDS = (
    "ibge_code",
    "municipality_name",
    "population_2022",
    "cattle_head_2024",
    "swine_head_2024",
    "poultry_head_2024",
    "snis_rdo_t_year",
    "snis_rpu_t_year",
    "snis_rdo_rpu_t_year",
)

TIMESERIES_SQL = """
INSERT INTO municipality_timeseries
    (ibge_code, year, source_id, variable, value, unit, quality)
VALUES
    (%(ibge_code)s, %(year)s, %(source_id)s, %(variable)s,
     %(value)s, %(unit)s, %(quality)s)
ON CONFLICT (ibge_code, year, source_id, variable) DO UPDATE
SET value = EXCLUDED.value, unit = EXCLUDED.unit, quality = EXCLUDED.quality
"""

POPULATION_SQL = """
UPDATE municipalities SET
    population = %(population)s,
    population_year = 2022,
    population_density = CASE
        WHEN area_km2 IS NULL OR area_km2 <= 0 THEN NULL
        ELSE ROUND((%(population)s::numeric / area_km2), 2)
    END
WHERE ibge_code = %(ibge_code)s
"""


def dsn() -> str:
    url = os.environ.get("DATABASE_URL")
    if url:
        return url
    return "postgresql://{u}:{p}@{h}:{port}/{db}".format(
        u=os.environ.get("POSTGRES_USER", "postgres"),
        p=os.environ.get("POSTGRES_PASSWORD", "password"),
        h=os.environ.get("POSTGRES_HOST", "db"),
        port=os.environ.get("POSTGRES_PORT", "5432"),
        db=os.environ.get("POSTGRES_DB", "cp2b_maps"),
    )


def _sidra_value(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "-":
        return 0.0
    if text in {"", "..", "...", "X", "None"}:
        return None
    return float(text.replace(".", "").replace(",", "."))


def _snapshot_value(value: Any) -> float | None:
    """Parse our normalized UTF-8 CSV, whose decimal separator is always '.'."""
    text = str(value).strip()
    return None if text in {"", "None"} else float(text)


def _fetch_json(url: str) -> list[dict[str, Any]]:
    with urllib.request.urlopen(url, timeout=120) as response:
        payload = json.load(response)
    if not isinstance(payload, list) or len(payload) < 2:
        raise ValueError(f"SIDRA returned no data for {url}")
    return payload[1:]  # first row is SIDRA's field dictionary


def _parse_snis(workbook: Path) -> dict[str, dict[str, float | None]]:
    """Read the official 2022 information table.

    Excel columns 75/80/85 are CO119/CO111/CO115 respectively. They are
    published as annual tonnes. The assertions make a changed workbook layout
    fail loudly instead of silently moving a daily or agent-level column into
    the annual field.
    """
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    try:
        ws = wb["Informacoes"]
        row8 = [cell.value for cell in next(ws.iter_rows(min_row=8, max_row=8))]
        for excel_col in (75, 80, 85):
            if str(row8[excel_col - 1]).strip() != "Total":
                raise ValueError(f"SNIS workbook layout changed at column {excel_col}")

        rows: dict[str, dict[str, float | None]] = {}
        for row in ws.iter_rows(min_row=10, values_only=True):
            if len(row) < 85 or row[3] not in {"MG", "SP"} or row[1] is None:
                continue
            code = str(row[1]).strip()
            rows[code] = {
                "snis_rdo_rpu_t_year": _sidra_value(row[74]),  # CO119
                "snis_rdo_t_year": _sidra_value(row[79]),  # CO111
                "snis_rpu_t_year": _sidra_value(row[84]),  # CO115
            }
        return rows
    finally:
        wb.close()


def build_snapshot(snis_workbook: Path, output: Path) -> list[dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}

    for uf in PUBLIC_STATE_COUNTS:
        for item in _fetch_json(PPM_URL.format(uf=uf)):
            code = str(item["D1C"])
            target = HERD_VARIABLES.get(str(item["D4N"]))
            if not target:
                continue
            rows.setdefault(
                code,
                {"ibge_code": code, "municipality_name": str(item["D1N"]).rsplit(" - ", 1)[0]},
            )[target] = _sidra_value(item["V"])

        for item in _fetch_json(CENSO_URL.format(uf=uf)):
            if str(item["D2C"]) != "93" or str(item["D3C"]) != str(YEAR_CENSO):
                continue
            code = str(item["D1C"])
            rows.setdefault(
                code,
                {"ibge_code": code, "municipality_name": str(item["D1N"]).rsplit(" - ", 1)[0]},
            )["population_2022"] = _sidra_value(item["V"])

    snis = _parse_snis(snis_workbook)
    for code, values in snis.items():
        if code in rows:
            rows[code].update(values)

    ordered = [rows[code] for code in sorted(rows)]
    validate_snapshot(ordered)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in ordered:
            writer.writerow({key: row.get(key) for key in FIELDS})

    metadata = {
        "scope": ["SP", "MG"],
        "municipalities": len(ordered),
        "sources": {
            "population": {
                "source": "IBGE Censo 2022",
                "sidra_table": 4709,
                "url_template": CENSO_URL,
            },
            "livestock": {"source": "IBGE PPM 2024", "sidra_table": 3939, "url_template": PPM_URL},
            "urban": {
                "source": "SNIS-RS 2022 Planilha de Informações",
                "url": SNIS_URL,
                "variables": {
                    "CO111": "RDO t/year",
                    "CO115": "RPU t/year",
                    "CO119": "RDO+RPU t/year",
                },
            },
        },
        "coverage": {
            "population": sum(row.get("population_2022") is not None for row in ordered),
            "ppm_primary_herds": sum(
                all(row.get(key) is not None for key in HERD_VARIABLES.values()) for row in ordered
            ),
            "snis_respondents": len(snis),
            "snis_co111_reporters": sum(row.get("snis_rdo_t_year") is not None for row in ordered),
        },
        "totals": {
            "population_2022": sum(float(row["population_2022"]) for row in ordered),
            "snis_co111_t_year": sum(float(row.get("snis_rdo_t_year") or 0) for row in ordered),
            "snis_co115_t_year": sum(float(row.get("snis_rpu_t_year") or 0) for row in ordered),
            "snis_co119_t_year": sum(float(row.get("snis_rdo_rpu_t_year") or 0) for row in ordered),
        },
    }
    output.with_suffix(".metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return ordered


def load_snapshot(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    numeric = set(FIELDS) - {"ibge_code", "municipality_name"}
    for row in rows:
        for key in numeric:
            row[key] = _snapshot_value(row.get(key))
    validate_snapshot(rows)
    return rows


def validate_snapshot(rows: list[dict[str, Any]]) -> None:
    codes = [str(row["ibge_code"]) for row in rows]
    if len(rows) != PUBLIC_MUNICIPALITIES or len(set(codes)) != PUBLIC_MUNICIPALITIES:
        raise ValueError(
            f"expected {PUBLIC_MUNICIPALITIES} unique public-state rows, got {len(set(codes))}"
        )
    if any(len(code) != 7 or code[:2] not in PUBLIC_STATE_COUNTS for code in codes):
        raise ValueError("snapshot contains an out-of-scope or malformed IBGE code")
    for uf, expected in PUBLIC_STATE_COUNTS.items():
        got = sum(code.startswith(uf) for code in codes)
        if got != expected:
            raise ValueError(f"UF {uf}: expected {expected} municipalities, got {got}")
    required = ("population_2022",)
    missing = [code for code, row in zip(codes, rows) if any(row.get(k) is None for k in required)]
    if missing:
        examples = [
            (code, [key for key in required if row.get(key) is None])
            for code, row in zip(codes, rows)
            if code in set(missing[:10])
        ]
        raise ValueError(
            f"population missing for {len(missing)} municipalities; examples={examples}"
        )
    partial_herds = [
        code
        for code, row in zip(codes, rows)
        if 0 < sum(row.get(key) is not None for key in HERD_VARIABLES.values()) < 3
    ]
    if partial_herds:
        raise ValueError(f"partial PPM primary-herd rows: {partial_herds[:10]}")


def promotion_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    populations: list[dict[str, Any]] = []
    series: list[dict[str, Any]] = []
    for row in rows:
        code = str(row["ibge_code"])
        populations.append({"ibge_code": code, "population": int(float(row["population_2022"]))})
        for variable, field in (
            ("bovino", "cattle_head_2024"),
            ("suino_total", "swine_head_2024"),
            ("galinaceos_total", "poultry_head_2024"),
        ):
            if row.get(field) is None:
                continue
            series.append(
                {
                    "ibge_code": code,
                    "year": YEAR_PPM,
                    "source_id": "ibge_ppm",
                    "variable": variable,
                    "value": row[field],
                    "unit": "head",
                    "quality": "measured",
                }
            )
        for variable, field in (
            ("rdo_coletado", "snis_rdo_t_year"),
            ("rpu_coletado", "snis_rpu_t_year"),
            ("rdo_rpu_coletado", "snis_rdo_rpu_t_year"),
        ):
            if row.get(field) is not None:
                series.append(
                    {
                        "ibge_code": code,
                        "year": YEAR_SNIS,
                        "source_id": "snis",
                        "variable": variable,
                        "value": row[field],
                        "unit": "t",
                        "quality": "measured",
                    }
                )
        series.append(
            {
                "ibge_code": code,
                "year": YEAR_CENSO,
                "source_id": "ibge_censo2022",
                "variable": "populacao_residente",
                "value": row["population_2022"],
                "unit": "inhabitants",
                "quality": "measured",
            }
        )
    return populations, series


def validate_database_spine(database_codes: set[str], snapshot_codes: set[str]) -> None:
    """Require the database and reviewed SP+MG snapshot to cover the same municipalities."""
    if database_codes == snapshot_codes:
        return
    missing = sorted(snapshot_codes - database_codes)
    extra = sorted(database_codes - snapshot_codes)
    raise ValueError(
        "SP + MG database spine does not match the reviewed snapshot; "
        f"missing={missing[:10]}, extra={extra[:10]}"
    )


def promote(rows: list[dict[str, Any]], dry_run: bool) -> None:
    populations, series = promotion_rows(rows)
    print(f"SP + MG snapshot: {len(rows)} municipalities, {len(series)} timeseries rows")
    if dry_run:
        print("dry run — nothing written")
        return
    import psycopg2
    from psycopg2.extras import execute_batch

    connection = psycopg2.connect(dsn())
    try:
        with connection, connection.cursor() as cursor:
            cursor.execute(
                "SELECT ibge_code FROM municipalities " "WHERE LEFT(ibge_code::text, 2) = ANY(%s)",
                (list(PUBLIC_STATE_COUNTS),),
            )
            database_codes = {str(row[0]) for row in cursor.fetchall()}
            snapshot_codes = {str(row["ibge_code"]) for row in rows}
            validate_database_spine(database_codes, snapshot_codes)
            execute_batch(cursor, POPULATION_SQL, populations, page_size=500)
            execute_batch(cursor, TIMESERIES_SQL, series, page_size=1000)
        print("committed — SP + MG population, PPM and SNIS activity promoted")
    finally:
        connection.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--snapshot", type=Path, default=DEFAULT_SNAPSHOT)
    parser.add_argument("--build-snapshot", action="store_true")
    parser.add_argument("--snis-workbook", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if args.build_snapshot:
        if not args.snis_workbook:
            raise SystemExit("--build-snapshot requires --snis-workbook")
        rows = build_snapshot(args.snis_workbook, args.snapshot)
        print(f"wrote reviewed snapshot: {args.snapshot}")
    else:
        rows = load_snapshot(args.snapshot)
    promote(rows, args.dry_run)
    return 0


if __name__ == "__main__":
    sys.exit(main())
