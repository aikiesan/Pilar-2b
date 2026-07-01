#!/usr/bin/env python3
"""
make_reference_review_xlsx.py — build a manual-review workbook for the reference corpus.

Input : a CSV export of scientific_references (columns: id, primary_residue, doi, url,
        external_url, title, authors, journal, year, reference_type, validation_status, notes).
Output: PILAR2b_referencias_revisao.xlsx with:
  - "References (399)" sheet: one row per reference, sorted by canonical feedstock.
      * current_url / current_doi (cleaned), suspect_doi_reuse (same DOI on >1 residue),
        needs_url, bmp_extracted (from notes), notes_excerpt
      * editable columns (yellow): CORRECTED_URL, CORRECTED_DOI, STATUS (dropdown), REVIEWER_NOTES
      * red rows = suspect DOI reuse (verify first); amber = missing URL
  - "Instructions" sheet: the manual review process.
Returned file is handed back to the maintainer; corrected sheet is reconciled into
references_unified.csv / the database afterwards. No DOI is ever invented.

Usage: python3 scripts/make_reference_review_xlsx.py <export.csv> [out.xlsx]
Requires: openpyxl
"""

import csv
import importlib.util
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
_u = importlib.util.spec_from_file_location("u", HERE / "unify_references.py")
U = importlib.util.module_from_spec(_u)
_u.loader.exec_module(U)  # reuse CODE_MAP

DOI_RE = re.compile(r"^10\.", re.I)
BMP_RE = re.compile(
    r"(\d{2,4}(?:[.,]\d+)?)\s*(?:±\s*[\d.,]+\s*)?(?:m?[lL]|Nm?[lL]|L)\s*(?:CH[₄4])?\s*/\s*(?:kg|g)\s*(?:VS|SV)",
    re.I,
)


def clean_doi(d):
    d = (d or "").strip()
    d = re.sub(r'[>"<)\s.]+$', "", d)
    d = re.sub(r"^https?://(dx\.)?doi\.org/", "", d)
    return d if DOI_RE.match(d) else ""


def main(src, out="PILAR2b_referencias_revisao.xlsx"):
    rows = list(csv.DictReader(open(src, encoding="utf-8")))
    doi_res = defaultdict(set)
    for r in rows:
        d = clean_doi(r.get("doi"))
        if d:
            doi_res[d].add((r.get("primary_residue") or "").strip())

    headers = [
        "id",
        "primary_residue",
        "feedstock_canonico",
        "title",
        "authors",
        "journal",
        "year",
        "reference_type",
        "validation_status",
        "current_url",
        "current_doi",
        "suspect_doi_reuse",
        "needs_url",
        "bmp_extracted",
        "notes_excerpt",
        "CORRECTED_URL",
        "CORRECTED_DOI",
        "STATUS",
        "REVIEWER_NOTES",
    ]
    edit_cols = {"CORRECTED_URL", "CORRECTED_DOI", "STATUS", "REVIEWER_NOTES"}

    wb = Workbook()
    ws = wb.active
    ws.title = "References (399)"
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B45309" if h in edit_cols else "1F4E2E")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    data = []
    for r in rows:
        feed = U.CODE_MAP.get(
            (r.get("primary_residue") or "").strip(), (r.get("primary_residue") or "").upper()
        )
        d = clean_doi(r.get("doi"))
        url = (r.get("url") or r.get("external_url") or "").strip()
        note = (r.get("notes") or "").replace("\n", " ")
        bmps = BMP_RE.findall(note)
        suspect = "Y" if (d and len(doi_res[d]) > 1) else ""
        needs = "Y" if not url else ""
        rid = int(r["id"]) if str(r.get("id", "")).isdigit() else 0
        data.append(
            (
                feed,
                rid,
                [
                    r.get("id"),
                    r.get("primary_residue"),
                    feed,
                    r.get("title"),
                    r.get("authors"),
                    r.get("journal"),
                    r.get("year"),
                    r.get("reference_type"),
                    r.get("validation_status"),
                    url,
                    d,
                    suspect,
                    needs,
                    (bmps[0] if bmps else ""),
                    note[:200],
                    "",
                    "",
                    "",
                    "",
                ],
            )
        )
    data.sort(key=lambda x: (x[0], x[1]))

    red = PatternFill("solid", fgColor="FDE2E2")
    amber = PatternFill("solid", fgColor="FEF3C7")
    edit = PatternFill("solid", fgColor="FFF7CC")
    for _feed, _id, row in data:
        ws.append(row)
        ridx = ws.max_row
        if row[11] == "Y":
            for c in range(1, len(headers) + 1):
                ws.cell(ridx, c).fill = red
        elif row[12] == "Y":
            for c in range(1, len(headers) + 1):
                ws.cell(ridx, c).fill = amber
        for c in (16, 17, 18, 19):
            ws.cell(ridx, c).fill = edit

    widths = {
        "A": 6,
        "B": 22,
        "C": 18,
        "D": 46,
        "E": 30,
        "F": 24,
        "G": 6,
        "H": 22,
        "I": 16,
        "J": 46,
        "K": 30,
        "L": 8,
        "M": 8,
        "N": 10,
        "O": 50,
        "P": 40,
        "Q": 28,
        "R": 12,
        "S": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    dv = DataValidation(
        type="list",
        formula1='"OK,FIXED_URL,FIXED_DOI,FIXED_BOTH,REMOVE,KEEP_NO_DOI"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"R2:R{ws.max_row}")

    ins = wb.create_sheet("Instructions")
    for line in [
        "PILAR-2b - Revisao manual de referencias",
        "",
        "Colunas amarelas = preencher: CORRECTED_URL, CORRECTED_DOI, STATUS (dropdown), REVIEWER_NOTES",
        "VERMELHO = mesmo DOI em >1 residuo (provavel DOI errado) -> verificar primeiro",
        "AMARELO  = sem URL -> encontrar URL direta do artigo",
        "Verificar DOI: abrir https://doi.org/<DOI> e conferir titulo/autores/ano.",
        "Nunca inventar DOI. Se nao achar: CORRECTED_DOI vazio, STATUS=KEEP_NO_DOI ou REMOVE.",
        "Ao terminar, devolver o .xlsx para reconciliacao em references_unified.csv / banco.",
    ]:
        ins.append([line])
    ins.column_dimensions["A"].width = 110
    ins["A1"].font = Font(bold=True, size=14)

    wb.save(out)
    print(
        f'rows={len(data)} suspect={sum(1 for d in data if d[2][11] == "Y")} '
        f'needs_url={sum(1 for d in data if d[2][12] == "Y")} -> {out}'
    )


if __name__ == "__main__":
    main(
        sys.argv[1] if len(sys.argv) > 1 else "scientific_references.csv",
        sys.argv[2] if len(sys.argv) > 2 else "PILAR2b_referencias_revisao.xlsx",
    )
